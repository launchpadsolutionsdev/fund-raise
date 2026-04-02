"""Parse department Excel files.

Each spreadsheet has three sheets: REPORT, RAW, INSTRUCTIONS.
We parse REPORT for summary/breakdown data and RAW for individual gifts.
"""

import openpyxl
import pandas as pd
from datetime import datetime


def parse_department_file(file_path, department):
    """Parse an Excel file for a given department and return structured data."""
    wb = openpyxl.load_workbook(file_path, data_only=True)

    # Validate sheets exist
    sheet_names_lower = [s.lower() for s in wb.sheetnames]
    if "report" not in sheet_names_lower:
        raise ValueError(f"Missing 'REPORT' sheet in {department} file")

    report_sheet_name = wb.sheetnames[sheet_names_lower.index("report")]
    report_ws = wb[report_sheet_name]

    result = {
        "summary": {},
        "gift_types": [],
        "sources": [],
        "funds": [],
        "raw_gifts": [],
    }

    # Parse REPORT sheet
    _parse_report_sheet(report_ws, department, result)

    # Parse RAW sheet
    if "raw" in sheet_names_lower:
        raw_sheet_name = wb.sheetnames[sheet_names_lower.index("raw")]
        _parse_raw_sheet(file_path, raw_sheet_name, department, result)

    wb.close()
    return result


def _safe_float(val):
    """Convert a value to float, handling percentages and None."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        val = val.strip().replace("$", "").replace(",", "").replace("%", "")
        try:
            return float(val)
        except ValueError:
            return None
    return None


def _safe_int(val):
    """Convert a value to int, handling None."""
    f = _safe_float(val)
    return int(f) if f is not None else None


def _parse_report_sheet(ws, department, result):
    """Scan the REPORT sheet for summary metrics and breakdowns."""
    summary = result["summary"]
    max_row = ws.max_row or 100
    max_col = ws.max_column or 10

    # Read all cell values into a list for scanning
    rows = []
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=False):
        rows.append(row)

    parsing_gift_types = False
    parsing_sources = False
    parsing_funds = False
    # For Events: third-party data in columns D-F
    parsing_tp_gift_types = False
    parsing_tp_funds = False

    for i, row in enumerate(rows):
        cell_a = row[0].value if row[0].value else ""
        cell_b = row[1].value if len(row) > 1 and row[1].value else None
        label = str(cell_a).strip().lower() if cell_a else ""

        # Summary metrics
        if label.startswith("total gifts"):
            summary["total_gifts"] = _safe_int(cell_b)
            # Events third party
            if department == "events" and len(row) > 4 and row[4].value is not None:
                summary["third_party_total_gifts"] = _safe_int(row[4].value)
        elif label.startswith("total amount") or label.startswith("total bequest"):
            summary["total_amount"] = _safe_float(cell_b)
            if department == "events" and len(row) > 4 and row[4].value is not None:
                summary["third_party_total_amount"] = _safe_float(row[4].value)
        elif "goal" in label and label.endswith("goal"):
            summary["goal"] = _safe_float(cell_b)
            if department == "events" and len(row) > 4 and row[4].value is not None:
                summary["third_party_goal"] = _safe_float(row[4].value)
        elif label == "% to goal":
            val = _safe_float(cell_b)
            summary["pct_to_goal"] = val
            if department == "events" and len(row) > 4 and row[4].value is not None:
                summary["third_party_pct_to_goal"] = _safe_float(row[4].value)
        elif label == "average legacy gift" and department == "legacy_giving":
            summary["avg_gift"] = _safe_float(cell_b)
        elif "new confirmed expectancies" in label and department == "legacy_giving":
            summary["new_expectancies"] = _safe_int(cell_b)
        elif "open estates" in label and department == "legacy_giving":
            summary["open_estates"] = _safe_int(cell_b)
        elif "recorded expectancies" in label and department == "legacy_giving":
            summary["recorded_expectancies"] = _safe_int(cell_b)

        # Section detection
        if label == "gift type" or label == "gift types":
            parsing_gift_types = True
            parsing_sources = False
            parsing_funds = False
            # For events, check if third party header is in column D
            if department == "events" and len(row) > 3:
                parsing_tp_gift_types = True
            continue
        elif label == "source" or label == "sources":
            parsing_gift_types = False
            parsing_sources = True
            parsing_funds = False
            continue
        elif "gift by fund" in label or "gifts by fund" in label or label == "fund" or label == "funds":
            parsing_gift_types = False
            parsing_sources = False
            parsing_funds = True
            if department == "events" and len(row) > 3:
                parsing_tp_funds = True
            continue

        # Parse breakdown rows
        if parsing_gift_types and label and label != "total":
            if cell_a and cell_b is not None and not label.startswith("gift type"):
                pct = _safe_float(row[2].value) if len(row) > 2 else None
                result["gift_types"].append({
                    "gift_type": str(cell_a).strip(),
                    "amount": _safe_int(cell_b),
                    "pct_of_gifts": pct,
                    "category": "primary",
                })
                # Third party gift types for events
                if parsing_tp_gift_types and len(row) > 4 and row[4].value is not None:
                    tp_pct = _safe_float(row[5].value) if len(row) > 5 else None
                    result["gift_types"].append({
                        "gift_type": str(cell_a).strip(),
                        "amount": _safe_int(row[4].value),
                        "pct_of_gifts": tp_pct,
                        "category": "third_party",
                    })
            elif not cell_a or label == "":
                parsing_gift_types = False
                parsing_tp_gift_types = False

        if parsing_sources and label and label != "total":
            if cell_a and cell_b is not None and not label.startswith("source"):
                pct = _safe_float(row[2].value) if len(row) > 2 else None
                result["sources"].append({
                    "source": str(cell_a).strip(),
                    "amount": _safe_int(cell_b),
                    "pct_of_gifts": pct,
                })
            elif not cell_a or label == "":
                parsing_sources = False

        if parsing_funds and label and label != "total" and label != "grand total":
            if cell_a and cell_b is not None and not label.startswith("fund") and "gift by fund" not in label and "gifts by fund" not in label:
                fund_entry = {
                    "fund_name": str(cell_a).strip(),
                    "amount": _safe_float(cell_b),
                    "pct_of_total": _safe_float(row[2].value) if len(row) > 2 else None,
                    "category": "primary",
                }
                # Extra columns for Annual Giving & Direct Mail
                if department in ("annual_giving", "direct_mail") and len(row) > 6:
                    fund_entry["onetime_count"] = _safe_int(row[3].value) if len(row) > 3 else None
                    fund_entry["recurring_count"] = _safe_int(row[4].value) if len(row) > 4 else None
                    fund_entry["online_count"] = _safe_int(row[5].value) if len(row) > 5 else None
                    fund_entry["mailed_in_count"] = _safe_int(row[6].value) if len(row) > 6 else None
                    fund_entry["total_count"] = _safe_int(row[7].value) if len(row) > 7 else None
                result["funds"].append(fund_entry)

                # Third party funds for events
                if parsing_tp_funds and len(row) > 4:
                    tp_col_d = row[3].value if len(row) > 3 else None
                    tp_col_e = row[4].value if len(row) > 4 else None
                    if tp_col_d is not None:
                        result["funds"].append({
                            "fund_name": str(cell_a).strip(),
                            "amount": _safe_float(tp_col_d),
                            "pct_of_total": _safe_float(tp_col_e) if tp_col_e else None,
                            "category": "third_party",
                        })
            elif not cell_a or label == "":
                parsing_funds = False
                parsing_tp_funds = False


def _parse_raw_sheet(file_path, sheet_name, department, result):
    """Parse the RAW sheet using pandas for efficiency."""
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name)
    except Exception:
        return

    if df.empty:
        return

    # Standardize column names
    cols = df.columns.tolist()
    raw_gifts = []

    for _, row in df.iterrows():
        gift = {
            "primary_addressee": _get_col(row, cols, 0),
            "appeal_id": _get_col(row, cols, 1),
            "split_amount": _safe_float(_get_col(row, cols, 2)),
            "fund_description": _get_col(row, cols, 3),
            "gift_id": _safe_int(_get_col(row, cols, 4)),
            "gift_type": _get_col(row, cols, 5),
            "gift_reference": _get_col(row, cols, 6),
            "gift_date": _parse_date(_get_col(row, cols, 7)),
            "extra_field": _get_col(row, cols, 8) if len(cols) > 8 else None,
        }
        raw_gifts.append(gift)

    result["raw_gifts"] = raw_gifts


def _get_col(row, cols, idx):
    """Get a column value by index, handling missing columns."""
    if idx >= len(cols):
        return None
    val = row.iloc[idx]
    if pd.isna(val):
        return None
    return str(val) if not isinstance(val, (int, float, datetime)) else val


def _parse_date(val):
    """Parse a date value from Excel."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, str):
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(val, fmt).date()
            except ValueError:
                continue
    return None
